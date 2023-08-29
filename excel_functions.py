from openpyxl import load_workbook


class Excel_Funtions:

    def __init__(self,file_name,sheet_name):
        self.file = file_name
        self.sheet = sheet_name

    def row_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_row
    
    def column_count(self):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        return sheet.max_column
    
    def read_data(self, row_number, column_number):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        cell_value = sheet.cell(row=row_number, column=column_number).value
        if cell_value == None:
            return ""
        else:
            return cell_value
    
    def write_data(self, row_number, column_number, data):
        workbook = load_workbook(self.file)
        sheet = workbook[self.sheet]
        sheet.cell(row=row_number, column=column_number).value = data
        workbook.save(self.file)
        